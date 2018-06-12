# -*- coding: utf-8 -*-
import openpyxl
import string
from string import maketrans

from openpyxl import Workbook
from openpyxl.chart import(
    PieChart,
    Reference
)
from openpyxl.chart.series import DataPoint       # 프로그램에 필요한 openpyxl 라이브러리 활성화


def find_word_data(data_list):        # 파라미터: 텍스트에서 등장한 단어를 정리한 list()
    word_count = 0
    only_once = 1
    used_word = list()
    most_frequent = tuple()                 # 분석을 마친 단어를 저장할 리스트 (반복 분석 X)
    length = my_list.__len__()
    for word in my_list:
        word_count = data_list.count(word)
        mark = 1
        for x in used_word:
            if word == x:
                mark = 0            # 단어가 used_word 에 존재하면 다시 출력 되지 않는다
                break

        if mark == 1:
            print("The word " + str(word) + " was used " + str(word_count) + " times.")
            used_word.append(word)  # 단어가 처음으로 분석된거면 문장 출력, used_word 에 추가

        if only_once == 1:
            most_frequent = (word, word_count)    # most_frequent 초기 활성화
            only_once = 0

        if word_count > most_frequent[1]:
            most_frequent = (word, word_count)    # most_frequent 갱신 (조건식)

    print("The total word count of the document is" + str(length))
    print("The most frequently used word is " + str(most_frequent[0]) + " with a word count of " + str(most_frequent[1]))

    return used_word            # 텍스트에서 출력된 단어를 중복없이 1개씩 갖고 있는 리스트()


def write_text_xlsx_file(used_list):
    file1 = open("WordAnalysis.txt", 'w')
    word_string = list()
    value_string = list()

    excel_file = Workbook()
    sheet = excel_file.active
    end_word = "A" + str(length)
    end_data = "B" + str(length)
    cell_range_word = sheet["A2": end_word]
    cell_range_data = sheet["B2": end_data]
    cell_range = sheet["A1": end_data]  # 엑셀 범위 지정
    sheet["A1"].value = "단어"
    sheet["B1"].value = "횟수"
    sheet['C1'].value = "많이 쓰인 단어"
    sheet['D1'].value = "횟수"
    marker = 2      # 셀번호

    for word in used_list:
        word_string.append(str(word))
        sheet["A" + str(marker)] = str(word)
        value_string.append(my_list.count(word))
        sheet["B"+str(marker)] = my_list.count(word)
        marker += 1
        file1.write(str(word) + "   :   " + str(my_list.count(word)) + "\n")   # 엑셀, 텍스트파일에 자료 이동

    row_num = 0
    for row in cell_range:
        if row[1].value > (length / 200):
            row_num += 1
            sheet.cell(row = row_num, column = 4, value = row[1].value)
            sheet.cell(row = row_num, column = 3, value = row[0].value)    # 파이 차트에서 쓸 값 정리

    pie = PieChart()
    labels = Reference(sheet, min_col = 3, min_row = 2, max_row = marker - 1)
    data = Reference(sheet, min_col = 4, min_row = 1, max_row = marker - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Significant words in article"
    sheet.add_chart(pie, "M1")
    excel_file.save("Word_AnalyticsFinal.xlsx")


if __name__ == '__main__':             # Main Program
    text_file = open("ExperimentalPython.txt", 'r')
    my_lists = text_file.read()
    my_list = my_lists.translate(None, string.punctuation)
    my_list = my_list.translate(None, "\n")
    my_list = my_list.translate(None, '"')
    my_list = my_list.translate(None, "'")
    my_list = my_list.lower()
    my_list = my_list.split(' ')
    length = len(my_list)           # 파일 (.txt) 에서 텍스트를 가지고 온 후 리스트에 각 단어를 쪼개서 넣는다

    a = find_word_data(my_list)         # 텍스트에서 등장하는 단어의 수, 가장 많이 등장하는 단어 분석

    write_text_xlsx_file(a)      # 텍스트 파일 생성



